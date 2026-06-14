"""Build the human-read audit workbook from audit_metadata.json.

Produces Glidearr_Audit_Checklist.xlsx with two sheets:
  • Dashboard  — live progress (COUNTIF formulas), per-stage rollup, legend.
  • Checklist  — every .py/.md in ORGANIC execution order with dropdown
                 check-off columns, security-flag highlighting, stage banding.

Read-only w.r.t. the codebase; just emits the .xlsx. Re-run after _audit_extract.py.
"""
import json
from pathlib import Path
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[3]
rows = json.loads((ROOT / "audit_metadata.json").read_text(encoding="utf-8"))

FONT = "Calibri"
# palette
INK      = "1F2937"
HEAD_BG  = "1F3A5F"
HEAD_FG  = "FFFFFF"
BAND_A   = "FFFFFF"
BAND_B   = "EEF3F8"
ENTRY_BG = "FDEBD0"   # cold-start / always
OPTIN_BG = "FCF3CF"
TOOL_BG  = "EAEDED"
GREEN    = "C6EFCE"; GREEN_FG = "1E7B34"
YELLOW   = "FFEB9C"; YELLOW_FG = "9C6500"
RED      = "FFC7CE"; RED_FG = "9C0006"
SEC_HI   = "F5B7B1"   # del/exec
SEC_MED  = "FAD7A0"   # net/sec
TITLE_BG = "0B2545"

thin = Side(style="thin", color="D5DBDB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# ── Checklist sheet ──────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Checklist"

COLS = [
    ("#",                              5),
    ("Stage",                         24),
    ("When / Run-phase",              20),
    ("Component",                     16),
    ("File (path from repo root)",    54),
    ("Type",                           6),
    ("Sec*",                           9),
    ("Purpose — what it CLAIMS to do",60),
    ("Top defs (classes · functions)",34),
    ("① Human-read?",                 15),
    ("② Does what it says?",          18),
    ("③ .md accurate?",               15),
    ("Reviewer",                      12),
    ("Date",                          12),
    ("Notes / issues found",          46),
    ("LOC",                            6),
    ("Paired .md",                    26),
    ("▶",                              4),
]
for i, (name, w) in enumerate(COLS, 1):
    c = ws.cell(row=1, column=i, value=name)
    c.font = Font(name=FONT, bold=True, color=HEAD_FG, size=10)
    c.fill = PatternFill("solid", fgColor=HEAD_BG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 30

# column index map
C = {name: idx for idx, (name, _) in enumerate(COLS, 1)}

# alternate band colour as the stage_code changes
band_toggle = {}
order_codes = []
for r in rows:
    if r["stage_code"] not in order_codes:
        order_codes.append(r["stage_code"])
for i, code in enumerate(order_codes):
    band_toggle[code] = BAND_A if i % 2 == 0 else BAND_B

excel_row = 2
for r in rows:
    band = band_toggle[r["stage_code"]]
    sec = r.get("sec", "")
    vals = [
        r["order"],
        r["stage"],
        r["phase"],
        r["component"],
        r["path"],
        r["ext"].lstrip(".").upper(),
        sec,
        r.get("doc", ""),
        r.get("defs", ""),
        "",   # ① read
        "",   # ② does
        "",   # ③ md
        "",   # reviewer
        "",   # date
        "",   # notes
        r.get("loc", ""),
        r.get("paired_md") or ("—" if r["ext"] == ".py" else ""),
        "▶" if r.get("runnable") else "",
    ]
    for ci, v in enumerate(vals, 1):
        cell = ws.cell(row=excel_row, column=ci, value=v)
        cell.font = Font(name=FONT, size=9, color=INK)
        cell.fill = PatternFill("solid", fgColor=band)
        cell.border = border
        if ci in (C["Purpose — what it CLAIMS to do"], C["Notes / issues found"],
                  C["Top defs (classes · functions)"]):
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        elif ci in (C["File (path from repo root)"], C["Paired .md"]):
            cell.alignment = Alignment(vertical="center", horizontal="left")
        else:
            cell.alignment = Alignment(vertical="center",
                                       horizontal="center" if ci in (C["#"], C["Type"], C["Sec*"], C["LOC"], C["▶"], C["When / Run-phase"]) else "left")
    # monospace-ish emphasis for the file path
    ws.cell(row=excel_row, column=C["File (path from repo root)"]).font = Font(name="Consolas", size=9, color=INK)
    excel_row += 1

LAST = excel_row - 1   # last data row

ws.freeze_panes = "F2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{LAST}"

# ── data-validation dropdowns ────────────────────────────────────────────────
dv_read = DataValidation(type="list",
    formula1='"✔ Pass,⚠ Concern,✖ Fail,— N/A"', allow_blank=True)
dv_does = DataValidation(type="list",
    formula1='"Yes,Partial,No,N/A"', allow_blank=True)
dv_md = DataValidation(type="list",
    formula1='"Yes,Stale,Wrong,No .md"', allow_blank=True)
for dv in (dv_read, dv_does, dv_md):
    ws.add_data_validation(dv)
col_read = get_column_letter(C["① Human-read?"])
col_does = get_column_letter(C["② Does what it says?"])
col_md   = get_column_letter(C["③ .md accurate?"])
dv_read.add(f"{col_read}2:{col_read}{LAST}")
dv_does.add(f"{col_does}2:{col_does}{LAST}")
dv_md.add(f"{col_md}2:{col_md}{LAST}")

# ── conditional formatting ───────────────────────────────────────────────────
def rng(colname):
    L = get_column_letter(C[colname])
    return f"{L}2:{L}{LAST}"

# security column: DEL/EXEC = red, NET/SEC = amber (DEL/EXEC rule first, stops)
sec_rng = rng("Sec*")
secL = get_column_letter(C["Sec*"])
ws.conditional_formatting.add(sec_rng,
    FormulaRule(formula=[f'OR(ISNUMBER(SEARCH("DEL",${secL}2)),ISNUMBER(SEARCH("EXEC",${secL}2)))'],
                fill=PatternFill("solid", fgColor=SEC_HI), stopIfTrue=True))
ws.conditional_formatting.add(sec_rng,
    FormulaRule(formula=[f'OR(ISNUMBER(SEARCH("NET",${secL}2)),ISNUMBER(SEARCH("SEC",${secL}2)))'],
                fill=PatternFill("solid", fgColor=SEC_MED)))

# ① read
ws.conditional_formatting.add(rng("① Human-read?"),
    CellIsRule(operator="equal", formula=['"✔ Pass"'], fill=PatternFill("solid", fgColor=GREEN), font=Font(color=GREEN_FG, bold=True)))
ws.conditional_formatting.add(rng("① Human-read?"),
    CellIsRule(operator="equal", formula=['"⚠ Concern"'], fill=PatternFill("solid", fgColor=YELLOW), font=Font(color=YELLOW_FG, bold=True)))
ws.conditional_formatting.add(rng("① Human-read?"),
    CellIsRule(operator="equal", formula=['"✖ Fail"'], fill=PatternFill("solid", fgColor=RED), font=Font(color=RED_FG, bold=True)))
# ② does
ws.conditional_formatting.add(rng("② Does what it says?"),
    CellIsRule(operator="equal", formula=['"Yes"'], fill=PatternFill("solid", fgColor=GREEN), font=Font(color=GREEN_FG)))
ws.conditional_formatting.add(rng("② Does what it says?"),
    CellIsRule(operator="equal", formula=['"Partial"'], fill=PatternFill("solid", fgColor=YELLOW), font=Font(color=YELLOW_FG)))
ws.conditional_formatting.add(rng("② Does what it says?"),
    CellIsRule(operator="equal", formula=['"No"'], fill=PatternFill("solid", fgColor=RED), font=Font(color=RED_FG, bold=True)))
# ③ md
ws.conditional_formatting.add(rng("③ .md accurate?"),
    CellIsRule(operator="equal", formula=['"Yes"'], fill=PatternFill("solid", fgColor=GREEN), font=Font(color=GREEN_FG)))
ws.conditional_formatting.add(rng("③ .md accurate?"),
    CellIsRule(operator="equal", formula=['"Stale"'], fill=PatternFill("solid", fgColor=YELLOW), font=Font(color=YELLOW_FG)))
ws.conditional_formatting.add(rng("③ .md accurate?"),
    CellIsRule(operator="equal", formula=['"Wrong"'], fill=PatternFill("solid", fgColor=RED), font=Font(color=RED_FG, bold=True)))

# ── Dashboard sheet ──────────────────────────────────────────────────────────
dash = wb.create_sheet("Dashboard", 0)
dash.sheet_view.showGridLines = False
for col, w in (("A", 3), ("B", 32), ("C", 12), ("D", 12), ("E", 12), ("F", 12), ("G", 14), ("H", 3), ("I", 40)):
    dash.column_dimensions[col].width = w

def dcell(ref, val, *, bold=False, size=10, color=INK, bg=None, align="left", wrap=False, italic=False):
    c = dash[ref]
    c.value = val
    c.font = Font(name=FONT, bold=bold, size=size, color=color, italic=italic)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    return c

dash.merge_cells("B2:G2")
dcell("B2", "Glidearr — Human-Read Audit Checklist", bold=True, size=16, color="FFFFFF", bg=TITLE_BG, align="left")
dash.row_dimensions[2].height = 30
dash.merge_cells("B3:G3")
dcell("B3", "Organic execution order · every .py + .md · verify each file does what it claims",
      italic=True, size=10, color="FFFFFF", bg=HEAD_BG)

CK = "Checklist!"
read_col = f"{CK}${col_read}$2:${col_read}${LAST}"
does_col = f"{CK}${col_does}$2:${col_does}${LAST}"
md_col   = f"{CK}${col_md}$2:${col_md}${LAST}"
file_col = f"{CK}$E$2:$E${LAST}"

# headline KPIs
dcell("B5", "OVERALL PROGRESS", bold=True, size=12, color=TITLE_BG)
kpi = [
    ("B6", "Total files to audit",        f"=COUNTA({file_col})"),
    ("B7", "① Human-read complete",       f'=COUNTIF({read_col},"✔ Pass")+COUNTIF({read_col},"⚠ Concern")+COUNTIF({read_col},"✖ Fail")+COUNTIF({read_col},"— N/A")'),
    ("B8", "   └ % reviewed",             None),
    ("B9", "✔ Passed clean",              f'=COUNTIF({read_col},"✔ Pass")'),
    ("B10","⚠ Concerns flagged",          f'=COUNTIF({read_col},"⚠ Concern")'),
    ("B11","✖ Failed read",               f'=COUNTIF({read_col},"✖ Fail")'),
    ("B12","② Confirmed does-what-it-says",f'=COUNTIF({does_col},"Yes")'),
    ("B13","   └ ⚠ Partial / ✖ No",       f'=COUNTIF({does_col},"Partial")+COUNTIF({does_col},"No")'),
    ("B14","③ .md docs accurate",         f'=COUNTIF({md_col},"Yes")'),
    ("B15","   └ Stale / Wrong .md",      f'=COUNTIF({md_col},"Stale")+COUNTIF({md_col},"Wrong")'),
]
for ref, label, formula in kpi:
    dcell(ref, label, bold=label.startswith(("①","②","③","Total")))
    if formula:
        vref = "C" + ref[1:]
        dcell(vref, formula, bold=True, align="center")
# % reviewed
dcell("C8", "=IF(C6=0,0,C7/C6)", bold=True, align="center", color="1E7B34")
dash["C8"].number_format = "0.0%"

# data-bar on the % cell
from openpyxl.formatting.rule import DataBarRule
dash.conditional_formatting.add("C8",
    DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1,
                color="2E86C1", showValue=True))

# ── per-stage rollup ──
dcell("B17", "PROGRESS BY EXECUTION STAGE", bold=True, size=12, color=TITLE_BG)
hdr = ["Stage (organic run order)", "Files", "Reviewed", "% Done", "Concerns", "Does-NOT", "Sec-flagged"]
for j, h in enumerate(hdr):
    dcell(f"{get_column_letter(2+j)}18", h, bold=True, color=HEAD_FG, bg=HEAD_BG,
          align="center", wrap=True)
dash.row_dimensions[18].height = 28

# distinct (code,label) in run order
seen = OrderedDict()
for r in rows:
    key = (r["stage_code"], r["stage"])
    seen.setdefault(key, 0)
    seen[key] += 1

stage_rng = f"{CK}$B$2:$B${LAST}"
sec_rngd  = f"{CK}$G$2:$G${LAST}"
rr = 19
for (code, label), n in seen.items():
    lbl = f'"{label}"'
    dcell(f"B{rr}", f"{code} · {label}")
    dcell(f"C{rr}", f'=COUNTIF({stage_rng},{lbl})', align="center")
    dcell(f"D{rr}", f'=COUNTIFS({stage_rng},{lbl},{read_col},"<>")', align="center")
    pct = dcell(f"E{rr}", f"=IF(C{rr}=0,0,D{rr}/C{rr})", align="center")
    pct.number_format = "0%"
    dcell(f"F{rr}", f'=COUNTIFS({stage_rng},{lbl},{read_col},"⚠ Concern")', align="center", color=YELLOW_FG)
    dcell(f"G{rr}", f'=COUNTIFS({stage_rng},{lbl},{does_col},"No")', align="center", color=RED_FG)
    # sec-flagged count for the stage (non-empty Sec col)
    dcell(f"H{rr}", None)  # spacer unused
    rr += 1
# Sec-flagged column actually belongs at col H header; recompute into a dedicated col
# (we placed header 'Sec-flagged' at col H index 8 -> letter H). Fill it:
rr = 19
for (code, label), n in seen.items():
    lbl = f'"{label}"'
    cc = dash[f"H{rr}"]
    cc.value = f'=COUNTIFS({stage_rng},{lbl},{sec_rngd},"?*")'
    cc.font = Font(name=FONT, size=10, color="9C0006")
    cc.alignment = Alignment(horizontal="center")
    rr += 1
TOTAL_ROW = rr
dcell(f"B{TOTAL_ROW}", "TOTAL", bold=True, bg=BAND_B)
dcell(f"C{TOTAL_ROW}", f"=SUM(C19:C{rr-1})", bold=True, align="center", bg=BAND_B)
dcell(f"D{TOTAL_ROW}", f"=SUM(D19:D{rr-1})", bold=True, align="center", bg=BAND_B)
pc = dcell(f"E{TOTAL_ROW}", f"=IF(C{TOTAL_ROW}=0,0,D{TOTAL_ROW}/C{TOTAL_ROW})", bold=True, align="center", bg=BAND_B)
pc.number_format = "0%"
dcell(f"F{TOTAL_ROW}", f"=SUM(F19:F{rr-1})", bold=True, align="center", bg=BAND_B, color=YELLOW_FG)
dcell(f"G{TOTAL_ROW}", f"=SUM(G19:G{rr-1})", bold=True, align="center", bg=BAND_B, color=RED_FG)
dcell(f"H{TOTAL_ROW}", f"=SUM(H19:H{rr-1})", bold=True, align="center", bg=BAND_B, color=RED_FG)

dash.conditional_formatting.add(f"E19:E{rr-1}",
    DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1,
                color="58D68D", showValue=True))

# ── legend ──
lg = TOTAL_ROW + 2
dcell(f"B{lg}", "HOW TO USE", bold=True, size=12, color=TITLE_BG)
legend = [
    "Work top-to-bottom on the Checklist tab — rows are in the order the code actually executes from `python scripts/main.py`, not folder order.",
    "Open each file, read it, then set the three dropdowns: ① Human-read?  ② Does what it says?  ③ .md accurate?  Add findings in Notes.",
    "Filter the Checklist (row-1 arrows) by Stage, When, or Sec* to audit one subsystem or all risk files at once.",
    "",
    "When / Run-phase:  ALWAYS = on every run · OPT-IN = only if a config flag is set · IMPORT = library called by others ·",
    "                   STANDALONE/DEBUG/TEST = never run by main.py (manual) · ▶ = file is a runnable entrypoint.",
    "",
    "Sec*  (heuristic touchpoints — scrutinise these hardest):",
    "   NET = makes outbound HTTP / builds API URLs      SEC = handles api keys / tokens / secrets",
    "   DEL = deletes media / unmonitors / removes files  EXEC = subprocess / eval / pickle / yaml.load   FS = writes to disk",
    "   (red Sec cell = DEL or EXEC present · amber = NET or SEC present)",
    "",
    "① ✔ Pass / ⚠ Concern / ✖ Fail   ② Yes / Partial / No / N/A   ③ Yes / Stale / Wrong / No .md",
    "Regenerate rows after code changes:  python scripts/support/tools/_audit_extract.py  then  _audit_build_xlsx.py",
]
for i, line in enumerate(legend):
    dcell(f"B{lg+1+i}", line, size=9, color=INK, wrap=False)
    dash.merge_cells(f"B{lg+1+i}:I{lg+1+i}")

# force the app to recalc all formulas when the file is opened, so the Dashboard
# shows live numbers in Excel/LibreOffice even before LibreOffice bakes them.
try:
    wb.calculation.fullCalcOnLoad = True
except Exception:
    from openpyxl.workbook.properties import CalcProperties
    wb.calculation = CalcProperties(fullCalcOnLoad=True)

wb.save(ROOT / "Glidearr_Audit_Checklist.xlsx")
print(f"WROTE Glidearr_Audit_Checklist.xlsx  ({LAST-1} data rows, {len(seen)} stages)")
